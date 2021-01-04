import PlayerManager
PM = PlayerManager
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook



def convertToExcel():
    wb = Workbook()
    matches = wb.create_sheet("matchedPlayers")
    queued = wb.create_sheet("queuedPlayers")
    players = wb.create_sheet("players")
    
    for matchedPair in PM.matchedPlayers:
        matches.append([str(matchedPair[0]), str(matchedPair[1])])
    for queuedPlayer in PM.queuedPlayers:
        queued.append([str(queuedPlayer)])
        
    players.append(PM.getPlayerHeadings())
    for playerID in PM.players:
        players.append([str(playerID)] + PM.players[playerID].getData())
    
    wb.save('PlayerData2/balances.xlsx')

def getDataFromExcel():
    wb = load_workbook(filename = 'PlayerData2/balances.xlsx')
    setPlayers(wb)
    setQueued(wb)
    setMatches(wb)
    
def setMatches(wb):
    queued = wb['queuedPlayers']
    for row in queued:
        PM.queuedPlayers.append(str(row[0].value))

def setQueued(wb):
    matches = wb['matchedPlayers']
    print(len(matches))
    for row in matches:
        PM.matchedPlayers.append([str(row[0].value), str(row[1].value)])

def setPlayers(wb):
    players = wb['players']
    
    headings = players[1]
    rows = list(players.rows)
    
    for row in rows[1:]:
        PID = str(getValueFromRow(row, headings, 'id'))
        name = str(getValueFromRow(row, headings, 'name'))
        elo = float(getValueFromRow(row, headings, 'elo'))
        gamesPlayed = int(getValueFromRow(row, headings, 'gamesPlayed'))
        gamesWon = int(getValueFromRow(row, headings, 'gamesWon'))
        lastOppentID = str(getValueFromRow(row, headings, 'lastOpponentID'))
        
        PM.players[PID] = PM.Player(PID, name)
        PM.players[PID].setData(elo, gamesPlayed, gamesWon, lastOppentID)

def getValueFromRow(row, headings,headingName):
    headingsList = []
    for cell in headings:
        headingsList += [cell.value]
    index = headingsList.index(headingName)
    return row[index].value

# =============================================================================
# PM.loadPlayers()
# convertToExcel()
# =============================================================================

# =============================================================================
# getDataFromExcel()
# =============================================================================















